from openpyxl import load_workbook
import csv

wb = load_workbook('./Копия ШР.xlsx')
sheet = wb['TDSheet']

titles_list = []

for i in range(1, 294):
  title_dict = {}
  if sheet.cell(row=i, column=2).value:

    title_dict['fio'] = sheet.cell(row=i, column=2).value
    title_dict['title'] = sheet.cell(row=i, column=6).value
    titles_list.append(title_dict)

fields = ['fio', 'title']

with open('./data.csv', 'w', encoding='cp1251') as csvfile:
  writer = csv.DictWriter(csvfile, fieldnames=fields)
  writer.writeheader()
  writer.writerows(titles_list)